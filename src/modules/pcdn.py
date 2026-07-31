#!/usr/bin/env python


import datetime
from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pywebio.output import put_button, put_file, put_loading, put_markdown, put_scope, put_text, use_scope
from pywebio.pin import pin, put_file_upload

from utils import display_random_pet


class Pcdn:
    def __init__(self):
        display_random_pet()

        put_markdown("# PCDN 统计")

        put_file_upload(
            name="pcdn_file",
            label="上传 PCDN文件",
            accept=".xlsx",
            placeholder="excel 格式的文件",
        )
        put_button(
            label="开始匹配文件",
            onclick=self.match_file,
        )

        put_markdown("----")
        put_scope("output")

    @use_scope(name="output", clear=True)
    def match_file(self):
        with use_scope(name="output", clear=True):
            try:
                with put_loading():
                    put_text("开始吭哧吭哧生成结果......")

                    uploaded_file = pin.pcdn_file
                    if uploaded_file is None:
                        put_text("请先上传文件")
                        return

                    file = BytesIO(uploaded_file["content"])

                    file.seek(0)
                    wb_src = load_workbook(file, data_only=True)
                    ws_src = wb_src.active
                    if ws_src is None:
                        return

                    merged_cells_row1 = []
                    merged_cells_row2 = []
                    for merged_range in ws_src.merged_cells.ranges:
                        if merged_range.min_row == 1 and merged_range.max_row == 1:
                            merged_cells_row1.append(merged_range)
                        elif merged_range.min_row == 2 and merged_range.max_row == 2:
                            merged_cells_row2.append(merged_range)

                    first_row_values = []
                    for col in range(1, ws_src.max_column + 1):
                        first_row_values.append(ws_src.cell(row=1, column=col).value)

                    second_row_values = []
                    for col in range(1, ws_src.max_column + 1):
                        second_row_values.append(ws_src.cell(row=2, column=col).value)

                    file.seek(0)
                    df = pd.read_excel(file, header=1)

                    df["日期"] = pd.to_datetime(df["日期"], errors="coerce")
                    df = df.dropna(subset=["日期", "账号"])
                    df = df.sort_values(by=["账号", "日期"], ascending=[True, True])
                    df_final = df.drop_duplicates(subset=["账号"], keep="last")

                    wb_out = Workbook()
                    ws_out = wb_out.active
                    if ws_out is None:
                        return

                    ws_out.title = "Sheet1"

                    for col_idx, value in enumerate(first_row_values, start=1):
                        ws_out.cell(row=1, column=col_idx, value=value)

                    for col_idx, col_name in enumerate(df_final.columns, start=1):
                        ws_out.cell(row=2, column=col_idx, value=col_name)

                    for merged_range in merged_cells_row1:
                        ws_out.merge_cells(f"{merged_range.coord}")

                    for merged_range in merged_cells_row2:
                        ws_out.merge_cells(f"{merged_range.coord}")

                    for row_idx, row_data in enumerate(df_final.values, start=3):
                        for col_idx, value in enumerate(row_data, start=1):
                            cell = ws_out.cell(row=row_idx, column=col_idx, value=value)
                            if df_final.columns[col_idx - 1] == "日期" and isinstance(value, (pd.Timestamp, datetime.datetime)):
                                cell.number_format = "YYYY-M-D"

                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    center_alignment = Alignment(horizontal="center", vertical="center")
                    bold_font = Font(bold=True)
                    thin_side = Side(style="thin", color="000000")
                    default_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

                    for row in ws_out.iter_rows(min_row=1, max_row=2):
                        for cell in row:
                            cell.fill = yellow_fill
                            cell.alignment = center_alignment
                            cell.font = bold_font
                            cell.border = default_border

                    for cell in ws_out[2]:
                        if cell.value:
                            width = sum(2.0 if ord(c) > 255 else 1.2 for c in str(cell.value))
                            ws_out.column_dimensions[get_column_letter(cell.column)].width = width + 2

                    ws_out.merge_cells("A1:A2")
                    ws_out.merge_cells("B1:B2")

                    ws_out.column_dimensions["A"].width = 12
                    ws_out.column_dimensions["B"].width = 14
                    ws_out.column_dimensions["C"].width = 16

                    output_buffer = BytesIO()
                    wb_out.save(output_buffer)
                    output_buffer.seek(0)
                    content = output_buffer.getvalue()

                put_file(
                    name="output.xlsx",
                    content=content,
                    label=">> 点击下载生成后的文件 <<",
                )

            except KeyError as e:
                put_text(f"文件缺少必要列: {e}")
            except Exception as e:
                put_text(f"运行时错误：\n{e}")


if __name__ == "__main__":
    Pcdn()
